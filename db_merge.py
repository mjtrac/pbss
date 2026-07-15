#!/usr/bin/env python3
"""
db_merge.py — bCounter multi-database results viewer and merger.

Opens a GUI where the user selects two or more counter_results.db files
(from different scanning machines), shows per-database vote totals side by
side, then merges them into a single output database.

Usage:
    python3 db_merge.py

Requirements: Python 3.9+, tkinter (included with Python on macOS).
"""

import sqlite3
import shutil
import os
import sys
import csv
import tempfile
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from collections import defaultdict
from pathlib import Path
from typing import Optional


# ── Database helpers ──────────────────────────────────────────────────────────

def get_vote_totals(db_path: str) -> dict:
    """
    Return {contest_title: {candidate_name: {"voted": N, "overvoted": N}}} 
    from a counter_results.db.
    """
    con = sqlite3.connect(db_path)
    rows = con.execute("""
        SELECT k.contest_title, c.candidate_name,
               SUM(CASE WHEN v.vote_status = 'VOTED'     THEN 1 ELSE 0 END),
               SUM(CASE WHEN v.vote_status = 'OVERVOTED' THEN 1 ELSE 0 END)
        FROM vote_opportunity v
        JOIN candidate c ON c.id = v.candidate_id_fk
        JOIN contest   k ON k.id = v.contest_id
        GROUP BY k.contest_title, c.candidate_name
        ORDER BY k.contest_title, c.candidate_name
    """).fetchall()
    total_ballots = con.execute("SELECT COUNT(*) FROM ballot_image").fetchone()[0]
    con.close()

    result = {"_total_ballots": total_ballots, "_contests": defaultdict(dict)}
    for contest, candidate, voted, overvoted in rows:
        result["_contests"][contest][candidate] = {
            "voted": voted or 0,
            "overvoted": overvoted or 0,
        }
    return result


def merge_databases(db_paths: list[str], out_path: str) -> tuple[int, list[str]]:
    """
    Merge multiple counter_results databases into out_path.
    Returns (total_ballots_merged, list_of_warnings).

    Strategy:
    - Copy the first DB as the base.
    - For each additional DB, INSERT all rows with offsetted PKs so they
      don't conflict.  Foreign keys are adjusted by the same offset.
    - Barcode and contest/candidate tables are deduplicated by value.
    """
    warnings = []
    shutil.copy2(db_paths[0], out_path)

    base = sqlite3.connect(out_path)
    base.execute("PRAGMA foreign_keys = OFF")
    base.execute("PRAGMA journal_mode = WAL")

    total_merged = base.execute("SELECT COUNT(*) FROM ballot_image").fetchone()[0]

    for idx, src_path in enumerate(db_paths[1:], start=2):
        src = sqlite3.connect(src_path)

        # ── Offset: push all PKs past the current max in the destination ──────
        max_ballot   = base.execute("SELECT COALESCE(MAX(id),0) FROM ballot_image").fetchone()[0]
        max_barcode  = base.execute("SELECT COALESCE(MAX(id),0) FROM barcode").fetchone()[0]
        max_contest  = base.execute("SELECT COALESCE(MAX(id),0) FROM contest").fetchone()[0]
        max_cand     = base.execute("SELECT COALESCE(MAX(id),0) FROM candidate").fetchone()[0]
        max_vote     = base.execute("SELECT COALESCE(MAX(id),0) FROM vote_opportunity").fetchone()[0]

        # ── Barcode — deduplicate by raw_data ─────────────────────────────────
        barcode_map: dict[int, int] = {}   # src_id → dest_id
        for row in src.execute("SELECT id, raw_data, jurisdiction_id, region_id, "
                               "party_id, ballot_type_id, election_id, page_number "
                               "FROM barcode"):
            src_id, raw_data = row[0], row[1]
            existing = base.execute(
                "SELECT id FROM barcode WHERE raw_data=?", (raw_data,)
            ).fetchone()
            if existing:
                barcode_map[src_id] = existing[0]
            else:
                new_id = max_barcode + src_id
                base.execute(
                    "INSERT INTO barcode (id,raw_data,jurisdiction_id,region_id,"
                    "party_id,ballot_type_id,election_id,page_number) VALUES (?,?,?,?,?,?,?,?)",
                    (new_id, raw_data, row[2], row[3], row[4], row[5], row[6], row[7])
                )
                barcode_map[src_id] = new_id

        # ── Contest — deduplicate by title + type ─────────────────────────────
        contest_map: dict[int, int] = {}
        for row in src.execute(
                "SELECT id, contest_title, contest_type, max_votes FROM contest"):
            src_id, title, ctype, maxv = row
            existing = base.execute(
                "SELECT id FROM contest WHERE contest_title=? AND contest_type=?",
                (title, ctype)
            ).fetchone()
            if existing:
                contest_map[src_id] = existing[0]
            else:
                new_id = max_contest + src_id
                base.execute(
                    "INSERT INTO contest (id, contest_title, contest_type, max_votes) "
                    "VALUES (?,?,?,?)",
                    (new_id, title, ctype, maxv)
                )
                contest_map[src_id] = new_id

        # ── Candidate — deduplicate by name + contest ─────────────────────────
        candidate_map: dict[int, int] = {}
        for row in src.execute(
                "SELECT id, contest_id, candidate_name, ballot_candidate_id, write_in "
                "FROM candidate"):
            src_id, src_contest_id, name, ballot_cid, write_in = row
            dest_contest_id = contest_map.get(src_contest_id, src_contest_id + max_contest)
            existing = base.execute(
                "SELECT id FROM candidate WHERE candidate_name=? AND contest_id=?",
                (name, dest_contest_id)
            ).fetchone()
            if existing:
                candidate_map[src_id] = existing[0]
            else:
                new_id = max_cand + src_id
                base.execute(
                    "INSERT INTO candidate (id, contest_id, candidate_name, "
                    "ballot_candidate_id, write_in) VALUES (?,?,?,?,?)",
                    (new_id, dest_contest_id, name, ballot_cid, write_in)
                )
                candidate_map[src_id] = new_id

        # ── Ballot image — check for duplicate image_path ─────────────────────
        ballot_map: dict[int, int] = {}
        skipped = 0
        for row in src.execute(
                "SELECT id, image_path, image_name, counted_at, dpi, page_number, "
                "was_rotated, corners_found, warp_dpi, canonical_width, canonical_height, "
                "corner_marks, barcode_id FROM ballot_image"):
            src_id     = row[0]
            image_path = row[1]
            src_bc_id  = row[12]
            existing   = base.execute(
                "SELECT id FROM ballot_image WHERE image_path=?", (image_path,)
            ).fetchone()
            if existing:
                ballot_map[src_id] = existing[0]
                skipped += 1
            else:
                new_id      = max_ballot + src_id
                dest_bc_id  = barcode_map.get(src_bc_id, src_bc_id + max_barcode)
                base.execute(
                    "INSERT INTO ballot_image (id, image_path, image_name, counted_at, "
                    "dpi, page_number, was_rotated, corners_found, warp_dpi, "
                    "canonical_width, canonical_height, corner_marks, barcode_id) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (new_id, image_path, row[2], row[3], row[4], row[5],
                     row[6], row[7], row[8], row[9], row[10], row[11], dest_bc_id)
                )
                ballot_map[src_id] = new_id
                total_merged += 1

        if skipped:
            warnings.append(f"DB {idx}: {skipped} duplicate image path(s) skipped")

        # ── Vote opportunity — remap all FKs ──────────────────────────────────
        vo_cols = ("id, ballot_image_id, contest_id, candidate_id_fk, abs_left, abs_top, "
                   "indicator_width, indicator_height, warp_dpi, image_x, image_y, "
                   "threshold, dark_pct, dark_pixels, total_pixels, mean_intensity, "
                   "vote_status")
        for row in src.execute(f"SELECT {vo_cols} FROM vote_opportunity"):
            (src_vo_id, src_bi_id, src_co_id, src_ca_id,
             abs_l, abs_t, ind_w, ind_h, wdpi, imgx, imgy,
             thresh, dpct, dpx, tpx, mint, vstatus) = row
            dest_bi = ballot_map.get(src_bi_id)
            dest_co = contest_map.get(src_co_id)
            dest_ca = candidate_map.get(src_ca_id)
            if dest_bi is None or dest_co is None or dest_ca is None:
                warnings.append(f"DB {idx}: skipped vote_opportunity {src_vo_id} "
                                 f"(unmapped FK)")
                continue
            new_vo_id = max_vote + src_vo_id
            base.execute(
                "INSERT INTO vote_opportunity "
                "(id, ballot_image_id, contest_id, candidate_id_fk, abs_left, abs_top, "
                "indicator_width, indicator_height, warp_dpi, image_x, image_y, "
                "threshold, dark_pct, dark_pixels, total_pixels, mean_intensity, "
                "vote_status) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (new_vo_id, dest_bi, dest_co, dest_ca,
                 abs_l, abs_t, ind_w, ind_h, wdpi, imgx, imgy,
                 thresh, dpct, dpx, tpx, mint, vstatus)
            )

        src.close()

    base.commit()
    base.execute("PRAGMA foreign_keys = ON")
    base.close()
    return total_merged, warnings




# ── Export helpers ────────────────────────────────────────────────────────────

def build_export_rows(dbs: list) -> list[dict]:
    """
    Build a flat list of rows for CSV/Excel export.

    Each row represents one (db, contest, candidate[, rank]) combination.
    After each group of per-db rows for a candidate, a SUM row is inserted
    followed by a blank row — so a user can manually verify the sum in a
    spreadsheet by selecting the per-db rows and summing them.

    Columns:
      Contest | Candidate | Rank (blank if non-ranked) |
      DB 1 Votes | DB 1 Overvotes | DB 2 Votes | ... | Sum Votes | Sum Overvotes
    """
    from collections import defaultdict
    import re

    rank_pat = re.compile(r'(.+?)\s*\(Rank\s*(\d+)\)$')

    # Collect all contests and candidates across all dbs
    all_contests: dict[str, set] = defaultdict(set)
    for db in dbs:
        for contest, cands in db.totals["_contests"].items():
            all_contests[contest].update(cands.keys())

    rows = []
    db_labels = [f"DB {i+1}" for i in range(len(dbs))]

    for contest in sorted(all_contests.keys()):
        cands = sorted(all_contests[contest])
        for cand in cands:
            m = rank_pat.match(cand)
            base_name = m.group(1).strip() if m else cand
            rank      = m.group(2) if m else ""

            # Per-DB detail rows
            sum_v = 0
            sum_ov = 0
            for i, db in enumerate(dbs):
                v  = db.totals["_contests"].get(contest, {}).get(cand, {})
                voted     = v.get("voted", 0)
                overvoted = v.get("overvoted", 0)
                sum_v  += voted
                sum_ov += overvoted
                row = {
                    "Row Type":   "Detail",
                    "Contest":    contest,
                    "Candidate":  base_name,
                    "Rank":       rank,
                    "Source":     f"DB {i+1}",
                    "Votes":      voted,
                    "Overvotes":  overvoted,
                }
                rows.append(row)

            # SUM row — labelled so the user can see the calculated total
            rows.append({
                "Row Type":   "SUM",
                "Contest":    contest,
                "Candidate":  base_name,
                "Rank":       rank,
                "Source":     f"SUM (DB 1–{len(dbs)})",
                "Votes":      sum_v,
                "Overvotes":  sum_ov,
            })
            # Blank row for manual verification
            rows.append({
                "Row Type":   "blank",
                "Contest":    "",
                "Candidate":  "",
                "Rank":       "",
                "Source":     "← manual check: select Detail rows above and sum",
                "Votes":      "",
                "Overvotes":  "",
            })

    return rows


def export_csv(dbs: list, out_path: str) -> None:
    """Write export rows to a CSV file."""
    rows = build_export_rows(dbs)
    headers = ["Row Type", "Contest", "Candidate", "Rank",
               "Source", "Votes", "Overvotes"]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def export_xlsx(dbs: list, out_path: str) -> None:
    """Write export rows to an .xlsx file with formatting."""
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    rows = build_export_rows(dbs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vote Totals"

    # ── Styles ────────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="0F172A")
    SUM_FILL  = PatternFill("solid", fgColor="1E3A5F")
    BLANK_FILL= PatternFill("solid", fgColor="1E293B")
    ALT_FILL  = PatternFill("solid", fgColor="111827")

    thin = Side(style="thin", color="334155")
    border = Border(bottom=thin)

    hdr_font  = Font(name="Arial", bold=True, color="E2E8F0", size=10)
    sum_font  = Font(name="Arial", bold=True, color="38BDF8", size=10)
    det_font  = Font(name="Arial", color="CBD5E1", size=10)
    blank_font= Font(name="Arial", color="475569", size=9, italic=True)
    num_fmt   = "#,##0"

    headers = ["Row Type", "Contest", "Candidate", "Rank",
               "Source", "Votes", "Overvotes"]
    col_widths = [10, 38, 32, 8, 26, 12, 12]

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font    = hdr_font
        cell.fill    = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    # Data rows
    alt = False
    for r_idx, row in enumerate(rows, start=2):
        rtype = row["Row Type"]

        if rtype == "blank":
            fill = BLANK_FILL
            font = blank_font
            num_font = blank_font
        elif rtype == "SUM":
            fill = SUM_FILL
            font = sum_font
            num_font = sum_font
            alt = not alt   # toggle after each candidate group
        else:
            fill = ALT_FILL if alt else PatternFill("solid", fgColor="0F172A")
            font = det_font
            num_font = det_font

        for col, key in enumerate(headers, 1):
            val = row.get(key, "")
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.fill = fill
            cell.font = font
            if key in ("Votes", "Overvotes") and isinstance(val, int):
                cell.number_format = num_fmt
                cell.font = num_font
                cell.alignment = Alignment(horizontal="right")
            elif key == "Source" and rtype == "blank":
                cell.alignment = Alignment(horizontal="left")
                cell.font = blank_font
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")

        # Bottom border on SUM rows
        if rtype == "SUM":
            for col in range(1, len(headers)+1):
                ws.cell(row=r_idx, column=col).border = border

    # Auto-filter
    ws.auto_filter.ref = f"A1:G1"

    wb.save(out_path)


# ── GUI ───────────────────────────────────────────────────────────────────────

DARK_BG   = "#0f172a"
PANEL_BG  = "#1e293b"
BORDER    = "#334155"
ACCENT    = "#38bdf8"
ACCENT2   = "#818cf8"
TEXT      = "#e2e8f0"
MUTED     = "#64748b"
GREEN     = "#22c55e"
YELLOW    = "#eab308"
DANGER    = "#ef4444"
FONT_MONO = ("Menlo", 11)
FONT_UI   = ("SF Pro Display", 11) if sys.platform == "darwin" else ("Segoe UI", 11)
FONT_HEAD = ("SF Pro Display", 13, "bold") if sys.platform == "darwin" else ("Segoe UI", 13, "bold")
FONT_TITLE= ("SF Pro Display", 16, "bold") if sys.platform == "darwin" else ("Segoe UI", 16, "bold")


class DbEntry:
    """Holds one loaded database."""
    def __init__(self, path: str, label: str, totals: dict):
        self.path   = path
        self.label  = label
        self.totals = totals


class MergeApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("bCounter — Multi-Database Results & Merge")
        self.configure(bg=DARK_BG)
        self.geometry("1200x800")
        self.minsize(800, 600)
        self.dbs: list[DbEntry] = []
        self._build_ui()

    def _build_ui(self):
        # ── Header bar ────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=PANEL_BG, pady=12, padx=20)
        hdr.pack(fill="x")
        tk.Label(hdr, text="bCounter  —  Multi-Database Results Merger",
                 bg=PANEL_BG, fg=TEXT, font=FONT_TITLE).pack(side="left")
        tk.Label(hdr, text="Compare and merge scanning results from multiple machines",
                 bg=PANEL_BG, fg=MUTED, font=FONT_UI).pack(side="left", padx=16)

        # ── Toolbar ───────────────────────────────────────────────────────────
        bar = tk.Frame(self, bg=DARK_BG, pady=8, padx=16)
        bar.pack(fill="x")
        self._btn(bar, "＋  Add Database", self._add_db, ACCENT).pack(side="left")
        self._btn(bar, "✕  Remove Selected", self._remove_db, DANGER).pack(side="left", padx=6)
        tk.Frame(bar, bg=DARK_BG, width=20).pack(side="left")
        self._btn(bar, "⇄  Merge All  →  Save", self._merge, ACCENT2).pack(side="left")
        tk.Frame(bar, bg=DARK_BG, width=20).pack(side="left")
        self._btn(bar, "⬇  Export CSV", self._export_csv, "#10b981").pack(side="left")
        self._btn(bar, "⬇  Export Excel", self._export_xlsx, "#f59e0b").pack(side="left", padx=6)

        # ── DB list ───────────────────────────────────────────────────────────
        list_frame = tk.Frame(self, bg=PANEL_BG, padx=12, pady=8)
        list_frame.pack(fill="x", padx=12, pady=(0,4))
        tk.Label(list_frame, text="Loaded Databases", bg=PANEL_BG, fg=MUTED,
                 font=FONT_UI).pack(anchor="w")
        self.db_listbox = tk.Listbox(list_frame, bg=DARK_BG, fg=TEXT,
                                     selectbackground=ACCENT, selectforeground=DARK_BG,
                                     font=FONT_MONO, height=4, borderwidth=0,
                                     highlightthickness=1, highlightcolor=BORDER,
                                     activestyle="none")
        self.db_listbox.pack(fill="x", pady=4)
        self.db_listbox.bind("<<ListboxSelect>>", self._on_db_select)

        # ── Results notebook ──────────────────────────────────────────────────
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Dark.TNotebook", background=DARK_BG, borderwidth=0)
        style.configure("Dark.TNotebook.Tab",
                        background=PANEL_BG, foreground=MUTED,
                        padding=[12, 6], font=FONT_UI)
        style.map("Dark.TNotebook.Tab",
                  background=[("selected", DARK_BG)],
                  foreground=[("selected", TEXT)])

        self.nb = ttk.Notebook(self, style="Dark.TNotebook")
        self.nb.pack(fill="both", expand=True, padx=12, pady=(0,12))

        self._overview_tab()
        self._log_tab()

    def _btn(self, parent, text, cmd, color):
        b = tk.Button(parent, text=text, command=cmd,
                      bg=color, fg=DARK_BG, activebackground=color,
                      font=(FONT_UI[0], FONT_UI[1], "bold"),
                      relief="flat", cursor="hand2",
                      padx=14, pady=6, borderwidth=0)
        return b

    def _overview_tab(self):
        frame = tk.Frame(self.nb, bg=DARK_BG)
        self.nb.add(frame, text="Results")

        # Scrollable canvas for results table
        self.results_canvas = tk.Canvas(frame, bg=DARK_BG, highlightthickness=0)
        vsb = tk.Scrollbar(frame, orient="vertical",
                           command=self.results_canvas.yview)
        hsb = tk.Scrollbar(frame, orient="horizontal",
                           command=self.results_canvas.xview)
        self.results_canvas.configure(yscrollcommand=vsb.set,
                                      xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.results_canvas.pack(fill="both", expand=True)

        self.results_inner = tk.Frame(self.results_canvas, bg=DARK_BG)
        self.results_canvas_win = self.results_canvas.create_window(
            (0, 0), window=self.results_inner, anchor="nw")
        self.results_inner.bind("<Configure>", self._on_inner_configure)
        self.results_canvas.bind("<Configure>", self._on_canvas_configure)

        # Mousewheel
        self.results_canvas.bind_all("<MouseWheel>",
            lambda e: self.results_canvas.yview_scroll(-1*(e.delta//120), "units"))

        self._show_empty_results()

    def _log_tab(self):
        frame = tk.Frame(self.nb, bg=DARK_BG)
        self.nb.add(frame, text="Merge Log")
        self.log_text = tk.Text(frame, bg=PANEL_BG, fg=TEXT, font=FONT_MONO,
                                 insertbackground=TEXT, relief="flat",
                                 state="disabled", wrap="none")
        sb = tk.Scrollbar(frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True, padx=8, pady=8)

    def _on_inner_configure(self, event):
        self.results_canvas.configure(
            scrollregion=self.results_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.results_canvas.itemconfig(
            self.results_canvas_win, width=event.width)

    # ── Database management ───────────────────────────────────────────────────

    def _add_db(self):
        paths = filedialog.askopenfilenames(
            title="Select counter_results.db file(s)",
            filetypes=[("SQLite databases", "*.db"), ("All files", "*.*")],
            initialdir=str(Path.home() / "pbss" / "bCounter")
        )
        for path in paths:
            if any(db.path == path for db in self.dbs):
                continue
            try:
                totals = get_vote_totals(path)
                label  = f"[{len(self.dbs)+1}] {Path(path).name}  ({Path(path).parent})"
                self.dbs.append(DbEntry(path, label, totals))
                self.db_listbox.insert("end", label)
                self._log(f"Loaded: {path}  "
                          f"({totals['_total_ballots']} ballot images)")
            except Exception as e:
                messagebox.showerror("Error", f"Could not load:\n{path}\n\n{e}")
        self._refresh_results()

    def _remove_db(self):
        sel = self.db_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        self.dbs.pop(idx)
        self.db_listbox.delete(idx)
        # Renumber labels
        self.db_listbox.delete(0, "end")
        for i, db in enumerate(self.dbs):
            db.label = f"[{i+1}] {Path(db.path).name}  ({Path(db.path).parent})"
            self.db_listbox.insert("end", db.label)
        self._refresh_results()

    def _on_db_select(self, event):
        pass  # future: highlight column

    # ── Results rendering ─────────────────────────────────────────────────────

    def _show_empty_results(self):
        for w in self.results_inner.winfo_children():
            w.destroy()
        tk.Label(self.results_inner,
                 text="Add two or more databases to compare results.",
                 bg=DARK_BG, fg=MUTED, font=FONT_UI,
                 pady=40).pack()

    def _refresh_results(self):
        for w in self.results_inner.winfo_children():
            w.destroy()

        if not self.dbs:
            self._show_empty_results()
            return

        # ── Collect all contests and candidates ───────────────────────────────
        all_contests: dict[str, set] = defaultdict(set)
        for db in self.dbs:
            for contest, cands in db.totals["_contests"].items():
                all_contests[contest].update(cands.keys())

        # ── Header row: DB labels ──────────────────────────────────────────────
        col_count = len(self.dbs) + 1  # +1 for merged total
        pad = dict(padx=6, pady=3)

        def hdr(parent, text, col, span=1, color=TEXT, bg=PANEL_BG):
            tk.Label(parent, text=text, bg=bg, fg=color,
                     font=FONT_HEAD if col == 0 else FONT_UI,
                     anchor="w" if col == 0 else "e",
                     width=28 if col == 0 else 14)\
              .grid(row=0, column=col, columnspan=span, sticky="ew", **pad)

        def cell(parent, text, row, col, color=TEXT, bg=DARK_BG, bold=False):
            f = (FONT_UI[0], FONT_UI[1], "bold") if bold else FONT_UI
            tk.Label(parent, text=text, bg=bg, fg=color, font=f,
                     anchor="w" if col == 0 else "e",
                     width=28 if col == 0 else 14)\
              .grid(row=row, column=col, sticky="ew", **pad)

        # Ballot totals bar
        totals_frame = tk.Frame(self.results_inner, bg=PANEL_BG,
                                padx=12, pady=8)
        totals_frame.pack(fill="x", padx=8, pady=(8,4))
        tk.Label(totals_frame, text="Ballot Images Scanned",
                 bg=PANEL_BG, fg=MUTED, font=FONT_UI).grid(
                     row=0, column=0, sticky="w", padx=6)
        merged_total = 0
        for i, db in enumerate(self.dbs):
            n = db.totals["_total_ballots"]
            merged_total += n
            tk.Label(totals_frame,
                     text=f"DB {i+1}: {n:,}",
                     bg=PANEL_BG, fg=ACCENT, font=FONT_HEAD)\
              .grid(row=0, column=i+1, padx=12)
        tk.Label(totals_frame,
                 text=f"Merged: {merged_total:,}",
                 bg=PANEL_BG, fg=ACCENT2, font=FONT_HEAD)\
          .grid(row=0, column=len(self.dbs)+1, padx=12)

        # Per-contest tables
        for contest in sorted(all_contests.keys()):
            cands = sorted(all_contests[contest])

            section = tk.Frame(self.results_inner, bg=DARK_BG)
            section.pack(fill="x", padx=8, pady=(10,0))

            # Contest title
            tk.Label(section, text=contest, bg=DARK_BG, fg=ACCENT,
                     font=FONT_HEAD, anchor="w", pady=4)\
              .grid(row=0, column=0, columnspan=col_count+1,
                    sticky="w", padx=6)

            # Separator
            tk.Frame(section, bg=BORDER, height=1)\
              .grid(row=1, column=0, columnspan=col_count+1,
                    sticky="ew", padx=6, pady=2)

            # Column headers
            tk.Label(section, text="Candidate", bg=PANEL_BG, fg=MUTED,
                     font=FONT_UI, anchor="w", width=28)\
              .grid(row=2, column=0, sticky="ew", **pad)
            for i, db in enumerate(self.dbs):
                tk.Label(section, text=f"DB {i+1}  Votes",
                         bg=PANEL_BG, fg=MUTED, font=FONT_UI,
                         anchor="e", width=14)\
                  .grid(row=2, column=i+1, sticky="ew", **pad)
            tk.Label(section, text="Merged  Votes",
                     bg=PANEL_BG, fg=ACCENT2, font=FONT_UI,
                     anchor="e", width=14)\
              .grid(row=2, column=len(self.dbs)+1, sticky="ew", **pad)

            # Candidate rows
            for r, cand in enumerate(cands, start=3):
                bg = DARK_BG if r % 2 == 0 else "#111827"
                is_rank = "(Rank" in cand

                # Candidate name
                tk.Label(section, text=cand, bg=bg, fg=MUTED if is_rank else TEXT,
                         font=FONT_UI, anchor="w", width=28)\
                  .grid(row=r, column=0, sticky="ew", **pad)

                merged_votes = 0
                merged_ov    = 0
                max_votes    = 0

                db_votes = []
                for db in self.dbs:
                    v = db.totals["_contests"].get(contest, {}).get(cand, {})
                    voted     = v.get("voted", 0)
                    overvoted = v.get("overvoted", 0)
                    db_votes.append((voted, overvoted))
                    merged_votes += voted
                    merged_ov    += overvoted
                    if voted > max_votes:
                        max_votes = voted

                for i, (voted, overvoted) in enumerate(db_votes):
                    label = str(voted)
                    if overvoted:
                        label += f"  ({overvoted}⚠)"
                    color = GREEN if voted == max_votes and voted > 0 else TEXT
                    tk.Label(section, text=label, bg=bg, fg=color,
                             font=FONT_UI, anchor="e", width=14)\
                      .grid(row=r, column=i+1, sticky="ew", **pad)

                # Merged column
                label = str(merged_votes)
                if merged_ov:
                    label += f"  ({merged_ov}⚠)"
                tk.Label(section, text=label, bg=bg, fg=ACCENT2,
                         font=(FONT_UI[0], FONT_UI[1], "bold"),
                         anchor="e", width=14)\
                  .grid(row=r, column=len(self.dbs)+1, sticky="ew", **pad)

        # Legend
        leg = tk.Frame(self.results_inner, bg=DARK_BG, pady=12)
        leg.pack(fill="x", padx=8)
        tk.Label(leg, text="⚠ = overvote count  |  ",
                 bg=DARK_BG, fg=MUTED, font=FONT_UI).pack(side="left", padx=6)
        tk.Label(leg,
                 text="Green = leading vote-getter in that contest across all databases",
                 bg=DARK_BG, fg=MUTED, font=FONT_UI).pack(side="left")

    # ── Merge ─────────────────────────────────────────────────────────────────

    def _merge(self):
        if len(self.dbs) < 2:
            messagebox.showwarning("Merge", "Add at least two databases to merge.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Save merged database as…",
            defaultextension=".db",
            filetypes=[("SQLite database", "*.db")],
            initialfile="counter_results_merged.db",
            initialdir=str(Path.home() / "pbss")
        )
        if not out_path:
            return

        self._log(f"\n{'─'*60}")
        self._log(f"Merging {len(self.dbs)} databases → {out_path}")
        try:
            paths = [db.path for db in self.dbs]
            total, warnings = merge_databases(paths, out_path)
            for w in warnings:
                self._log(f"  ⚠  {w}")
            self._log(f"  ✓  Merged {total:,} ballot images total")
            self._log(f"  ✓  Output: {out_path}")
            messagebox.showinfo(
                "Merge Complete",
                f"Merged {total:,} ballot images from {len(self.dbs)} databases.\n\n"
                f"Saved to:\n{out_path}"
                + (f"\n\n{len(warnings)} warning(s) — see Merge Log tab." if warnings else "")
            )
            self.nb.select(1)  # switch to log tab
        except Exception as e:
            self._log(f"  ✗  Error: {e}")
            messagebox.showerror("Merge Failed", str(e))

    def _export_csv(self):
        if not self.dbs:
            messagebox.showwarning("Export", "Add at least one database first.")
            return
        path = filedialog.asksaveasfilename(
            title="Save CSV export as…",
            defaultextension=".csv",
            filetypes=[("CSV file", "*.csv")],
            initialfile="vote_totals.csv",
            initialdir=str(Path.home() / "pbss")
        )
        if not path:
            return
        try:
            export_csv(self.dbs, path)
            self._log(f"CSV exported: {path}")
            messagebox.showinfo("Exported", f"CSV saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    def _export_xlsx(self):
        if not self.dbs:
            messagebox.showwarning("Export", "Add at least one database first.")
            return
        path = filedialog.asksaveasfilename(
            title="Save Excel export as…",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile="vote_totals.xlsx",
            initialdir=str(Path.home() / "pbss")
        )
        if not path:
            return
        try:
            export_xlsx(self.dbs, path)
            self._log(f"Excel exported: {path}")
            messagebox.showinfo("Exported", f"Excel saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    def _log(self, text: str):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = MergeApp()
    app.mainloop()
